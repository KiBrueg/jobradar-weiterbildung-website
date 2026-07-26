from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def register_ceo_routes(app, db_path: Path, report_dir: Path, init_db_func):
    def con() -> sqlite3.Connection:
        c = sqlite3.connect(db_path)
        c.row_factory = sqlite3.Row
        return c

    def rows(sql: str, params: tuple = ()) -> list[dict[str, Any]]:
        with con() as c:
            return [dict(r) for r in c.execute(sql, params).fetchall()]

    def one(sql: str, params: tuple = ()) -> dict[str, Any] | None:
        with con() as c:
            r = c.execute(sql, params).fetchone()
            return dict(r) if r else None

    def ensure_metric_tables() -> None:
        init_db_func()
        with con() as c:
            c.executescript(
                """
                CREATE TABLE IF NOT EXISTS source_runs (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  workflow_run_id INTEGER DEFAULT 0,
                  source_name TEXT NOT NULL,
                  source_type TEXT DEFAULT 'source',
                  status TEXT DEFAULT 'completed',
                  raw_items INTEGER DEFAULT 0,
                  normalized_items INTEGER DEFAULT 0,
                  duplicate_items INTEGER DEFAULT 0,
                  relevant_items INTEGER DEFAULT 0,
                  failed_items INTEGER DEFAULT 0,
                  success_rate REAL DEFAULT 100,
                  quality_score REAL DEFAULT 0,
                  risk_level TEXT DEFAULT 'low',
                  started_at TEXT NOT NULL,
                  finished_at TEXT NOT NULL,
                  error_message TEXT DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS cost_events (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  tool_name TEXT NOT NULL,
                  tool_type TEXT DEFAULT 'tool',
                  source_name TEXT DEFAULT '',
                  workflow_run_id INTEGER DEFAULT 0,
                  units REAL DEFAULT 0,
                  unit_type TEXT DEFAULT 'request',
                  unit_cost REAL DEFAULT 0,
                  estimated_cost REAL DEFAULT 0,
                  created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS reports (
                  id INTEGER PRIMARY KEY AUTOINCREMENT,
                  school_id INTEGER REFERENCES schools(id) ON DELETE SET NULL,
                  course_id INTEGER REFERENCES courses(id) ON DELETE SET NULL,
                  report_type TEXT DEFAULT 'CEO Report',
                  period_start TEXT DEFAULT '',
                  period_end TEXT DEFAULT '',
                  status TEXT DEFAULT 'prepared',
                  file_path TEXT DEFAULT '',
                  created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS dashboard_settings (
                  key TEXT PRIMARY KEY,
                  value TEXT NOT NULL
                );
                """
            )
            defaults = {
                "manual_minutes_per_raw_job": "6",
                "qa_minutes_per_relevant_lead": "2",
                "coach_hourly_rate": "45",
                "report_value_estimate": "91",
                "target_qa_backlog": "10",
                "monthly_tool_budget": "180",
                "manual_jobs_per_coach_hour": "5",
            }
            c.executemany("INSERT OR IGNORE INTO dashboard_settings(key,value) VALUES(?,?)", list(defaults.items()))
            # Keep productivity math consistent with manual_minutes_per_raw_job=6: 10 checked jobs/hour.
            c.execute("UPDATE dashboard_settings SET value='10' WHERE key='manual_jobs_per_coach_hour' AND CAST(value AS REAL) < 10")
            if c.execute("SELECT COUNT(*) FROM source_runs").fetchone()[0] > 0:
                c.commit()
                return

            course_ids = [r[0] for r in c.execute("SELECT id FROM courses ORDER BY id").fetchall()] or [1]
            base = datetime.utcnow().replace(hour=8, minute=0, second=0, microsecond=0)
            sources = [
                ("SerpAPI Google Jobs", "Job Search API", 0.019, 96, 34, "low"),
                ("OpenRouter LLM", "LLM Parsing", 0.028, 92, 48, "low"),
                ("Jina Reader", "Page Fetch / CRAG", 0.006, 88, 22, "low"),
                ("Arbeitnow", "Free Source", 0.0, 94, 30, "low"),
                ("Remotive", "Free Source", 0.0, 86, 26, "medium"),
                ("RemoteOK", "Scraper/API", 0.0, 72, 20, "medium"),
                ("Jobicy", "Free Source", 0.0, 58, 12, "high"),
                ("Himalayas", "Free Source", 0.0, 67, 14, "medium"),
                ("Bundesagentur Jobsuche", "Public API", 0.0, 90, 28, "low"),
            ]
            wf_names = ["Connection Test", "Flow 4 Job APIs", "Weekly Course Radar", "Source Health Check"]
            for day_back in range(0, 365, 2):
                dt = base - timedelta(days=day_back)
                source = sources[day_back % len(sources)]
                raw = max(4, int(source[4] + (day_back % 9) - 4 + (dt.month % 4)))
                failed = 1 if day_back % 17 == 0 else 0
                duplicates = max(0, int(raw * (0.18 + (day_back % 4) * 0.02)))
                normalized = max(0, raw - failed)
                relevant = max(1, int((normalized - duplicates) * (source[3] / 100) * 0.38))
                status = "partial" if failed else "completed"
                created = dt.isoformat() + "Z"
                wf_id = c.execute(
                    "INSERT INTO workflow_runs(workflow_name,run_mode,status,input_count,output_count,note,created_at) VALUES(?,?,?,?,?,?,?)",
                    (wf_names[day_back % len(wf_names)], "test" if day_back % 5 == 0 else "prod", status, raw, relevant, "seeded CEO metric run", created),
                ).lastrowid
                c.execute(
                    """INSERT INTO source_runs(workflow_run_id,source_name,source_type,status,raw_items,normalized_items,duplicate_items,relevant_items,failed_items,success_rate,quality_score,risk_level,started_at,finished_at,error_message)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (wf_id, source[0], source[1], status, raw, normalized, duplicates, relevant, failed, max(0, source[3]-failed*8), min(98, source[3] - (duplicates/max(1, raw)*10)), source[5], created, (dt+timedelta(minutes=7)).isoformat()+"Z", "" if not failed else "temporary source issue"),
                )
                if source[2] > 0:
                    c.execute(
                        "INSERT INTO cost_events(tool_name,tool_type,source_name,workflow_run_id,units,unit_type,unit_cost,estimated_cost,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                        (source[0], source[1], source[0], wf_id, raw, "item/request", source[2], round(raw*source[2], 4), created),
                    )
                if day_back % 6 == 0:
                    tool = "Localtunnel" if day_back % 12 == 0 else "SQLite Local DB"
                    c.execute(
                        "INSERT INTO cost_events(tool_name,tool_type,source_name,workflow_run_id,units,unit_type,unit_cost,estimated_cost,created_at) VALUES(?,?,?,?,?,?,?,?,?)",
                        (tool, "Infrastructure", tool, wf_id, 1, "run", 0, 0, created),
                    )
                if day_back % 4 == 0:
                    cid = course_ids[(day_back // 4) % len(course_ids)]
                    score = max(45, min(96, 82 - (day_back % 29) + (cid % 4) * 4))
                    st = ["Candidate", "QA Needed", "Approved", "Rejected"][day_back % 4]
                    c.execute(
                        "INSERT INTO leads(course_id,title,provider,status,score,cost,why_fit,missing_evidence,risks,sources,email_draft,updated_at) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",
                        (cid, f"KPI Lead {day_back}", source[0], st, score, "metric", "Generated from historical source run for KPI cockpit.", "QA check needed." if st == "QA Needed" else "", "Senioritätsrisiko" if score < 65 else "", f"https://example.com/kpi/{day_back}", "", created),
                    )
                if day_back % 28 == 0:
                    cid = course_ids[(day_back // 28) % len(course_ids)]
                    c.execute(
                        "INSERT INTO reports(school_id,course_id,report_type,period_start,period_end,status,file_path,created_at) VALUES((SELECT school_id FROM courses WHERE id=?),?,?,?,?,?,?,?)",
                        (cid, cid, "CEO Monatsreport", (dt-timedelta(days=30)).date().isoformat(), dt.date().isoformat(), "prepared", "", created),
                    )
            c.commit()

    def setting(settings: dict[str, str], key: str, default: float) -> float:
        try:
            return float(settings.get(key, default))
        except Exception:
            return default

    def bounds(range_name: str) -> tuple[datetime, datetime, str]:
        end = datetime.utcnow()
        rn = (range_name or "monat").lower()
        if rn == "tag":
            return end.replace(hour=0, minute=0, second=0, microsecond=0), end, "Tagesreport"
        if rn == "woche":
            return end - timedelta(days=7), end, "Wochenreport"
        if rn == "jahr":
            return end.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0), end, "Jahresreport"
        return end.replace(day=1, hour=0, minute=0, second=0, microsecond=0), end, "Monatsreport"

    def clamp(v: float, lo: float = 0, hi: float = 100) -> float:
        return max(lo, min(hi, v))

    def aggregate(start: datetime, end: datetime, settings: dict[str, str]) -> dict[str, Any]:
        ss, es = start.isoformat()+"Z", end.isoformat()+"Z"
        sr = one("""SELECT COALESCE(SUM(raw_items),0) raw, COALESCE(SUM(normalized_items),0) normalized, COALESCE(SUM(duplicate_items),0) duplicates, COALESCE(SUM(relevant_items),0) relevant, COALESCE(SUM(failed_items),0) failed, COALESCE(AVG(success_rate),0) success_rate, COALESCE(AVG(quality_score),0) source_quality FROM source_runs WHERE started_at>=? AND started_at<?""", (ss, es)) or {}
        wr = one("SELECT COUNT(*) total, SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) completed FROM workflow_runs WHERE created_at>=? AND created_at<?", (ss, es)) or {}
        lc = one("""SELECT COUNT(*) leads, COALESCE(AVG(score),0) avg_score, SUM(CASE WHEN status IN ('QA Needed','Candidate','New') THEN 1 ELSE 0 END) qa_open, SUM(CASE WHEN lower(risks) LIKE '%senior%' OR lower(risks) LIKE '%risk%' THEN 1 ELSE 0 END) risky FROM leads WHERE updated_at>=? AND updated_at<?""", (ss, es)) or {}
        rc = one("SELECT COUNT(*) reports FROM reports WHERE created_at>=? AND created_at<?", (ss, es)) or {"reports": 0}
        costs = one("SELECT COALESCE(SUM(estimated_cost),0) total FROM cost_events WHERE created_at>=? AND created_at<?", (ss, es)) or {"total": 0}
        raw = float(sr.get("raw") or 0)
        relevant = float(sr.get("relevant") or 0)
        manual_h = raw * setting(settings, "manual_minutes_per_raw_job", 6) / 60
        auto_h = relevant * setting(settings, "qa_minutes_per_relevant_lead", 2) / 60
        saved = max(0, manual_h - auto_h)
        saved_pct = saved / manual_h * 100 if manual_h else 0
        gross = saved * setting(settings, "coach_hourly_rate", 45)
        report_val = float(rc.get("reports") or 0) * setting(settings, "report_value_estimate", 91)
        tool = float(costs.get("total") or 0)
        benefit = gross + report_val
        net = benefit - tool
        roi = benefit / tool if tool else 0
        total_runs = float(wr.get("total") or 0)
        completed = float(wr.get("completed") or 0)
        safety = completed / total_runs * 100 if total_runs else 100
        leads = float(lc.get("leads") or 0)
        avg = float(lc.get("avg_score") or 0)
        qa = float(lc.get("qa_open") or 0)
        risky = float(lc.get("risky") or 0)
        target = setting(settings, "target_qa_backlog", 10)
        fit = clamp(avg - ((risky/leads*10) if leads else 0) - max(0, qa-target)*0.5)
        source_q = float(sr.get("source_quality") or 0)
        budget = setting(settings, "monthly_tool_budget", 180)
        cost_control = clamp(100 - max(0, tool/budget*100-60))
        qa_health = clamp(100 - max(0, qa-target)*4)
        market = clamp(.35*fit + .25*safety + .20*source_q + .10*cost_control + .10*qa_health)
        manual_jobs_per_hour = setting(settings, "manual_jobs_per_coach_hour", 10)
        jobradar_items_per_review_hour = raw / auto_h if auto_h else 0
        productivity = jobradar_items_per_review_hour / manual_jobs_per_hour if manual_jobs_per_hour else 0
        return {"raw_jobs": int(raw), "normalized_items": int(sr.get("normalized") or 0), "duplicates": int(sr.get("duplicates") or 0), "relevant_leads": int(relevant), "failed_items": int(sr.get("failed") or 0), "created_leads": int(leads), "qa_open": int(qa), "reports": int(rc.get("reports") or 0), "manual_hours": round(manual_h, 1), "automated_hours": round(auto_h, 1), "time_saved_hours": round(saved, 1), "time_saved_percent": round(saved_pct), "gross_savings": round(gross, 2), "report_value": round(report_val, 2), "total_benefit": round(benefit, 2), "tool_costs": round(tool, 2), "net_benefit": round(net, 2), "roi": round(roi, 1), "automation_safety": round(safety), "arbeitsmarkt_fit": round(fit), "productivity_factor": round(productivity, 1), "manual_jobs_per_hour": round(manual_jobs_per_hour, 1), "jobradar_items_per_review_hour": round(jobradar_items_per_review_hour, 1), "source_quality": round(source_q), "cost_control": round(cost_control), "qa_health": round(qa_health), "market_index": round(market), "workflow_runs": int(total_runs), "successful_runs": int(completed)}

    def series(settings: dict[str, str]) -> list[dict[str, Any]]:
        now_dt = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        out = []
        for i in range(11, -1, -1):
            m = (now_dt - timedelta(days=i*31)).replace(day=1)
            n = (m + timedelta(days=32)).replace(day=1)
            a = aggregate(m, n, settings)
            a.update({"label": m.strftime("%b"), "month": m.strftime("%Y-%m")})
            out.append(a)
        return out

    @app.get("/api/ceo-dashboard")
    def ceo_dashboard(range: str = "monat") -> dict[str, Any]:
        ensure_metric_tables()
        start, end, label = bounds(range)
        settings = {r["key"]: r["value"] for r in rows("SELECT key,value FROM dashboard_settings")}
        current = aggregate(start, end, settings)
        span = end - start
        previous = aggregate(start - span, start, settings)
        ss, es = start.isoformat()+"Z", end.isoformat()+"Z"
        cost_rows = rows("SELECT tool_name,tool_type,source_name,COUNT(*) runs,ROUND(COALESCE(SUM(units),0),1) units,ROUND(COALESCE(SUM(estimated_cost),0),2) cost FROM cost_events WHERE created_at>=? AND created_at<? GROUP BY tool_name,tool_type,source_name ORDER BY cost DESC,runs DESC", (ss, es))
        source_health = rows("SELECT source_name,source_type,COUNT(*) runs,COALESCE(SUM(raw_items),0) raw_items,COALESCE(SUM(relevant_items),0) relevant_items,COALESCE(SUM(failed_items),0) failed_items,ROUND(AVG(success_rate),1) success_rate,ROUND(AVG(quality_score),1) quality_score,MAX(risk_level) risk_level,MAX(finished_at) last_run FROM source_runs WHERE started_at>=? AND started_at<? GROUP BY source_name,source_type ORDER BY success_rate DESC", (ss, es))
        course_perf = rows("""SELECT c.id course_id,c.name course,COALESCE(s.name,c.provider) school,c.fit_score base_fit,COUNT(l.id) leads,ROUND(COALESCE(AVG(l.score),0),1) avg_score,SUM(CASE WHEN l.status IN ('QA Needed','Candidate','New') THEN 1 ELSE 0 END) qa_open FROM courses c LEFT JOIN schools s ON s.id=c.school_id LEFT JOIN leads l ON l.course_id=c.id AND l.updated_at>=? AND l.updated_at<? GROUP BY c.id ORDER BY avg_score DESC,leads DESC""", (ss, es))
        for c in course_perf:
            val = float(c.get("avg_score") or c.get("base_fit") or 0)
            fit = round(clamp(val - max(0, (c.get("qa_open") or 0)-3)*1.5))
            c["arbeitsmarkt_fit"] = fit
            c["risk"] = "High" if fit < 60 else "Medium" if fit < 75 else "Low"
            c["recommendation"] = "Suchprofil überarbeiten" if fit < 60 else "Queries verfeinern" if fit < 75 else "Weiter skalieren"
        actions = [
            {"title":"Data Analyst & BI Kurse weiter skalieren","priority":"Hoch","impact":"+€600 Monatsnutzen möglich"},
            {"title":"Python Backend Suchprofil prüfen — Fit unter 60%","priority":"Hoch","impact":"Risiko senken"},
            {"title":"SerpAPI Queries bündeln","priority":"Mittel","impact":"API-Kosten stabil halten"},
            {"title":"QA-Rückstau reduzieren","priority":"Mittel","impact":"Reportqualität erhöhen"},
            {"title":"OpenRouter-Guthaben prüfen & auffüllen (Schwelle: $5)","priority":"Mittel","impact":"Workflow-Ausfälle verhindern"},
        ]
        ai_roles = [
            {"id":"ai_integration","role":"AI Integration Engineer","de":"KI-Integrationsspezialist","titles":["AI Integration Engineer","AI Automation Engineer","Workflow Automation Engineer","API Integration Engineer","n8n Automation Specialist","Low-Code Automation Specialist"],"exclude":["Senior","Lead","Principal","Head of","10+ years"],"courses":["AI Automation / n8n Course","Python Backend Bootcamp","Business Intelligence Track"],"fit":88,"leads":18,"entry_level":72,"risk":"Low","recommendation":"Als Standard-Zielrolle für Automatisierungs-Kurse nutzen"},
            {"id":"ai_agents","role":"AI Agent Engineer","de":"KI-Agentenentwickler","titles":["AI Agent Engineer","LLM Agent Developer","LangGraph Developer","Autonomous Agent Engineer","AI Workflow Agent Engineer"],"exclude":["Research Scientist","PhD","Principal","Robotics"],"courses":["AI Automation / n8n Course","Python Backend Bootcamp"],"fit":82,"leads":11,"entry_level":55,"risk":"Medium","recommendation":"Gut als Premium-/Aspirationsrolle, Junior-Filter streng halten"},
            {"id":"rag_knowledge","role":"RAG / AI Knowledge Engineer","de":"KI Knowledge Engineer / RAG Engineer","titles":["RAG Engineer","AI Knowledge Engineer","Knowledge Base Engineer","LLM Retrieval Engineer","Vector Database Engineer"],"exclude":["Senior","Staff","Research","PhD"],"courses":["Data Analyst Weiterbildung","Business Intelligence Track","AI Automation / n8n Course"],"fit":84,"leads":14,"entry_level":62,"risk":"Medium","recommendation":"Für Kurse mit Dokumenten-/Datenbank-Fokus aufnehmen"},
            {"id":"ai_data","role":"AI Data Engineer","de":"KI Data Engineer","titles":["AI Data Engineer","Data Pipeline Engineer","Data Engineer AI","Analytics Engineer","ETL Developer AI"],"exclude":["Senior","Lead","Cloud Architect","Big Data Architect"],"courses":["Data Analyst Weiterbildung","Business Intelligence Track","Python Backend Bootcamp"],"fit":86,"leads":21,"entry_level":68,"risk":"Low","recommendation":"Stark für Data/BI-Kurse und Arbeitsmarktberichte"},
            {"id":"llm_apps","role":"LLM Application Engineer","de":"LLM Anwendungsentwickler","titles":["LLM Engineer","LLM Application Engineer","AI Application Engineer","Prompt Engineer","LLM Workflow Engineer"],"exclude":["Fine-tuning Research","ML Research Scientist","PhD"],"courses":["AI Automation / n8n Course","Python Backend Bootcamp"],"fit":79,"leads":16,"entry_level":58,"risk":"Medium","recommendation":"Mit API-/Backend-Skills kombinieren, nicht nur Prompting"},
            {"id":"solutions_architect","role":"AI Solutions Architect","de":"KI Solution Architect","titles":["AI Solutions Architect","AI Consultant","AI Automation Consultant","AI Transformation Consultant"],"exclude":["Enterprise Architect 10+","Director","Head of"],"courses":["AI Automation / n8n Course","Business Intelligence Track"],"fit":71,"leads":8,"entry_level":35,"risk":"High","recommendation":"Nicht als Junior-Default; eher für Pitch und Upskilling-Benchmark"},
        ]
        return {"ok": True, "range": range, "period_label": label, "period_start": ss, "period_end": es, "settings": settings, "current": current, "previous": previous, "series": series(settings), "cost_breakdown": cost_rows, "source_health": source_health, "course_performance": course_perf, "ai_role_benchmark": ai_roles, "actions": actions}

    @app.get("/download/ceo-report.json")
    def download_ceo_report(range: str = "monat") -> FileResponse:
        data = ceo_dashboard(range)
        filename = f"JobRadar_CEO_Report_{range}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        report_dir.mkdir(exist_ok=True)
        path = report_dir / filename
        path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        return FileResponse(path, filename=filename, media_type="application/json")

    def _excel_value(value: Any) -> Any:
        if isinstance(value, list):
            return ", ".join(str(item) for item in value)
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)
        return value

    def _write_sheet(ws, rows_data: list[dict[str, Any]]) -> None:
        if not rows_data:
            ws.append(["Keine Daten"])
            return
        headers = list(rows_data[0].keys())
        ws.append(headers)
        for row in rows_data:
            ws.append([_excel_value(row.get(h, "")) for h in headers])
        header_fill = PatternFill("solid", fgColor="EAF2FF")
        thin = Side(style="thin", color="D9E2EC")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F2937")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        for col_idx, header in enumerate(headers, 1):
            width = min(42, max(len(str(header)) + 2, 12, *(len(str(ws.cell(r, col_idx).value or "")) + 2 for r in range(2, min(ws.max_row, 50) + 1))))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    @app.get("/download/ceo-report.xlsx")
    def download_ceo_report_xlsx(range: str = "monat") -> FileResponse:
        data = ceo_dashboard(range)
        filename = f"JobRadar_CEO_Report_{range}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_dir.mkdir(exist_ok=True)
        path = report_dir / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws["A1"] = "JobRadar CEO Report"
        ws["A1"].font = Font(bold=True, size=16, color="111827")
        ws["A2"] = "Zeitraum"
        ws["B2"] = data.get("period_label", range)
        ws["A3"] = "Erstellt"
        ws["B3"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
        ws["A5"] = "KPI"
        ws["B5"] = "Wert"
        for c in ws[5]:
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor="EAF2FF")
        current = data.get("current", {})
        summary_rows = [
            ("Rohstellen geprüft", current.get("raw_jobs", 0)),
            ("Relevante Leads", current.get("relevant_leads", 0)),
            ("Zeitersparnis %", current.get("time_saved_percent", 0)),
            ("Gesparte Stunden", current.get("time_saved_hours", 0)),
            ("Netto-Effekt EUR", current.get("net_benefit", 0)),
            ("ROI x", current.get("roi", 0)),
            ("Produktivitätsfaktor", current.get("productivity_factor", 0)),
            ("Tool-Kosten EUR", current.get("tool_costs", 0)),
            ("Arbeitsmarkt-Fit %", current.get("arbeitsmarkt_fit", 0)),
            ("Automation Safety %", current.get("automation_safety", 0)),
        ]
        for row_idx, (label, value) in enumerate(summary_rows, 6):
            ws.cell(row_idx, 1, label)
            ws.cell(row_idx, 2, value)
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

        sheets = [
            ("Kosten", data.get("cost_breakdown", [])),
            ("Quellen", data.get("source_health", [])),
            ("Kurse", data.get("course_performance", [])),
            ("AI Rollen", data.get("ai_role_benchmark", [])),
            ("Aktionen", data.get("actions", [])),
        ]
        for title, rows_data in sheets:
            _write_sheet(wb.create_sheet(title), rows_data)
        wb.save(path)
        return FileResponse(path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
